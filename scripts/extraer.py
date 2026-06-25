import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ============================================================
# CONFIGURACIÓN - Ajusta estas rutas según tu ambiente
# ============================================================
archivo_origen = r"D:\Users\Usuario\Desktop\QuickBooks\Facturación 2024.xlsx"
archivo_destino = r"D:\Users\Usuario\Desktop\QuickBooks\clientes_empresas_unicos.xlsx"

# ============================================================
# PROCESAMIENTO
# ============================================================
print("=" * 60)
print("📖 EXTRACCIÓN DE CLIENTES Y EMPRESAS ÚNICOS")
print("=" * 60)

# Leer el Excel
print("\n📂 Leyendo archivo...")
df = pd.read_excel(archivo_origen)
print(f"   Total filas encontradas: {len(df)}")

# Nombres de columnas (con saltos de línea como están en el Excel)
col_cliente = 'Cliente\n(OP)'
col_empresa = 'Compañía en la que se facturo'

# Verificar que existen las columnas
if col_cliente not in df.columns:
    print(f"❌ No se encontró columna: {col_cliente}")
    print("Columnas disponibles:")
    for c in df.columns:
        print(f"   - {repr(c)}")
    exit(1)

if col_empresa not in df.columns:
    print(f"❌ No se encontró columna: {col_empresa}")
    exit(1)

# Extraer solo esas dos columnas
df_filtrado = df[[col_cliente, col_empresa]].copy()
df_filtrado.columns = ['Cliente', 'Empresa']
df_filtrado = df_filtrado[df_filtrado['Cliente'].notna()]

# Para cada cliente, obtener la empresa (si tiene en alguna fila)
def obtener_empresa(cliente):
    filas = df_filtrado[df_filtrado['Cliente'] == cliente]
    empresas = filas['Empresa'].dropna().unique()
    if len(empresas) > 0:
        return empresas[0]
    return None

# Obtener clientes únicos
clientes_unicos = df_filtrado['Cliente'].unique()

# Crear DataFrame con clientes únicos y su empresa
data = []
for cliente in clientes_unicos:
    empresa = obtener_empresa(cliente)
    data.append({'Cliente': cliente, 'Empresa': empresa})

df_unicos = pd.DataFrame(data)

# Separar en dos grupos: CON empresa y SIN empresa
con_empresa = df_unicos[df_unicos['Empresa'].notna()].copy()
sin_empresa = df_unicos[df_unicos['Empresa'].isna()].copy()

# Ordenar: CON empresa por Empresa y luego por Cliente
con_empresa = con_empresa.sort_values(by=['Empresa', 'Cliente'])

# SIN empresa ordenar por Cliente
sin_empresa = sin_empresa.sort_values(by='Cliente')

# Concatenar: PRIMERO los que tienen empresa (agrupados), LUEGO los que no tienen
df_final = pd.concat([con_empresa, sin_empresa], ignore_index=True)

# Estadísticas
total = len(df_final)
num_con_empresa = len(con_empresa)
num_sin_empresa = len(sin_empresa)

print(f"\n📊 ESTADÍSTICAS:")
print(f"   Total clientes únicos: {total}")
print(f"   ✅ Con empresa asignada: {num_con_empresa}")
print(f"   ⚠️  Sin empresa (vacíos): {num_sin_empresa}")

# Mostrar empresas encontradas
print(f"\n🏢 EMPRESAS ENCONTRADAS:")
empresas_unicas = con_empresa['Empresa'].unique()
for e in sorted(empresas_unicas):
    count = (df_final['Empresa'] == e).sum()
    print(f"   - {e}: {count} clientes")

# ============================================================
# CREAR EXCEL CON FORMATO BONITO
# ============================================================
print("\n🎨 Aplicando formato al Excel...")

# Guardar primero sin formato
df_final.to_excel(archivo_destino, index=False)

# Abrir y aplicar formato
wb = load_workbook(archivo_destino)
ws = wb.active

# Estilos
header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Colores por empresa
empresa_colors = {
    'SX': PatternFill(start_color="D5E8D4", end_color="D5E8D4", fill_type="solid"),  # Verde claro
    'Laitcorp': PatternFill(start_color="DAE8FC", end_color="DAE8FC", fill_type="solid"),  # Azul claro
    'HW': PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),  # Naranja claro
    'Corp': PatternFill(start_color="E1D5E7", end_color="E1D5E7", fill_type="solid"),  # Morado claro
}
sin_empresa_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo claro

# Aplicar formato al encabezado
for col in range(1, 3):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = border

# Aplicar formato a las filas de datos
for row in range(2, len(df_final) + 2):
    empresa_value = ws.cell(row=row, column=2).value
    
    for col in range(1, 3):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(vertical='center')
        
        # Color según empresa
        if empresa_value is None or (isinstance(empresa_value, float) and pd.isna(empresa_value)):
            cell.fill = sin_empresa_fill
        elif empresa_value in empresa_colors:
            cell.fill = empresa_colors[empresa_value]

# Ajustar ancho de columnas
ws.column_dimensions['A'].width = 45
ws.column_dimensions['B'].width = 25

# Guardar
wb.save(archivo_destino)

print(f"\n✅ ARCHIVO GENERADO: {archivo_destino}")
print("=" * 60)
print("\n📝 LEYENDA DE COLORES:")
print("   🟢 Verde claro: SX (Soportexperto)")
print("   🔵 Azul claro: Laitcorp")
print("   🟠 Naranja claro: HW (Hardware y Network)")
print("   🟣 Morado claro: Corp (Corporación)")
print("   🟡 Amarillo: SIN EMPRESA - Completar")
print("=" * 60)
print("\n📝 SIGUIENTE PASO:")
print("   La operadora debe completar la columna 'Empresa'")
print("   para los clientes en amarillo (al final del archivo).")
print("=" * 60)