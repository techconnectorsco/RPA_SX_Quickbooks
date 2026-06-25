from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName

# ============================================================
# CONFIGURACIÓN
# ============================================================
archivo = r"D:\Users\Usuario\Desktop\QuickBooks\Facturación Nuevo.xlsx"
MAX_FILAS = 10000  # Rango de validación

print("=" * 60)
print("🔧 APLICANDO VALIDACIONES Y FORMATO")
print("=" * 60)

# Cargar el archivo
print("\n📂 Cargando archivo...")
wb = load_workbook(archivo)
ws = wb.active

# Obtener última columna
max_col = ws.max_column
last_col_letter = get_column_letter(max_col)

print(f"   Filas con datos: {ws.max_row}")
print(f"   Rango de validación: hasta fila {MAX_FILAS}")

# ============================================================
# 0. CREAR HOJA "Listas" PARA VALORES DE DESPLEGABLES
# ============================================================
print("\n📋 Creando hoja 'Listas' para desplegables...")

# Eliminar si ya existe
if "Listas" in wb.sheetnames:
    del wb["Listas"]

# Crear hoja
ws_listas = wb.create_sheet("Listas")

# Valores de STATUS (columna A)
status_valores = ["Ejecutar", "Pendiente", "Revisión", "Lista"]
for i, val in enumerate(status_valores, start=1):
    ws_listas.cell(row=i, column=1, value=val)

# Valores de EMPRESAS (columna B)
empresa_valores = [
    "Soportexperto.com S.A.",
    "Hardware y Network S.A.",
    "Corporación Latinoamericana de Tecnología T.I. S.A."
]
for i, val in enumerate(empresa_valores, start=1):
    ws_listas.cell(row=i, column=2, value=val)

# Encabezados para la hoja Listas
ws_listas.cell(row=1, column=1).font = Font(bold=True)
ws_listas.cell(row=1, column=2).font = Font(bold=True)
ws_listas.column_dimensions['A'].width = 15
ws_listas.column_dimensions['B'].width = 50

# Ocultar la hoja (cambiar a 'visible' si quieres verla)
ws_listas.sheet_state = 'hidden'

# ============================================================
# 1. LISTA DESPLEGABLE PARA STATUS (Columna A)
# ============================================================
print("\n📋 Aplicando lista desplegable en STATUS (columna A)...")

dv_status = DataValidation(
    type="list",
    formula1="=Listas!$A$1:$A$4",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Valor no válido",
    error="Solo se permiten: Ejecutar, Pendiente, Revisión, Lista (o vacío)"
)
dv_status.prompt = "Seleccione un status"
dv_status.promptTitle = "Status"

dv_status.add(f"A2:A{MAX_FILAS}")
ws.add_data_validation(dv_status)

# ============================================================
# 2. FORMATO CONDICIONAL PARA STATUS (toda la fila)
# ============================================================
print("🎨 Aplicando formato condicional para STATUS...")

verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
naranja = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

# LISTA = verde
ws.conditional_formatting.add(
    f"A2:{last_col_letter}{MAX_FILAS}",
    FormulaRule(formula=['$A2="Lista"'], fill=verde)
)

# PENDIENTE = amarillo
ws.conditional_formatting.add(
    f"A2:{last_col_letter}{MAX_FILAS}",
    FormulaRule(formula=['$A2="Pendiente"'], fill=amarillo)
)

# REVISIÓN = naranja
ws.conditional_formatting.add(
    f"A2:{last_col_letter}{MAX_FILAS}",
    FormulaRule(formula=['$A2="Revisión"'], fill=naranja)
)

# ============================================================
# 3. VALIDACIÓN DE FECHAS (Columnas K y L)
# ============================================================
print("📅 Aplicando validación de fechas (columnas K y L)...")

dv_fecha = DataValidation(
    type="date",
    operator="greaterThan",
    formula1="2020-01-01",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Fecha no válida",
    error="Ingrese una fecha válida (después de 01/01/2020). No escriba texto."
)
dv_fecha.prompt = "Ingrese una fecha válida"
dv_fecha.promptTitle = "Fecha"

dv_fecha.add(f"K2:K{MAX_FILAS}")
dv_fecha.add(f"L2:L{MAX_FILAS}")
ws.add_data_validation(dv_fecha)

# ============================================================
# 4. LISTA DESPLEGABLE PARA COMPAÑÍA (Columna F)
# ============================================================
print("🏢 Aplicando lista desplegable en Compañía (columna F)...")

dv_empresa = DataValidation(
    type="list",
    formula1="=Listas!$B$1:$B$3",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Empresa no válida",
    error="Seleccione una empresa de la lista"
)
dv_empresa.prompt = "Seleccione la empresa"
dv_empresa.promptTitle = "Compañía"

dv_empresa.add(f"F2:F{MAX_FILAS}")
ws.add_data_validation(dv_empresa)

# ============================================================
# 5. FORMATO DEL ENCABEZADO
# ============================================================
print("🎨 Aplicando formato al encabezado...")

header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=10)
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

for col in range(1, max_col + 1):
    cell = ws.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = border

# Ajustar anchos
ws.column_dimensions['A'].width = 12
ws.column_dimensions['F'].width = 45

# ============================================================
# 6. GUARDAR
# ============================================================
wb.save(archivo)

print(f"\n✅ ARCHIVO ACTUALIZADO: {archivo}")
print("=" * 60)
print("\n📝 RESUMEN DE VALIDACIONES APLICADAS:")
print(f"   📊 Rango: Filas 2 a {MAX_FILAS}")
print("")
print("   ✅ Columna A (STATUS): Lista desplegable")
print("      - Ejecutar (sin color) → Robot lo procesa")
print("      - Pendiente (amarillo) → Se salta")
print("      - Revisión (naranja) → Se salta")  
print("      - Lista (verde) → Ya procesada")
print("")
print("   ✅ Columna F (Compañía): Lista desplegable")
print("      - Soportexperto.com S.A.")
print("      - Hardware y Network S.A.")
print("      - Corporación Latinoamericana de Tecnología T.I. S.A.")
print("")
print("   ✅ Columnas K y L (Fechas): Solo fechas válidas")
print("")
print("   📋 Hoja 'Listas' creada (oculta)")
print("=" * 60)
print("\n💡 PARA AGREGAR EMPRESAS EN EL FUTURO:")
print("   1. Click derecho en pestaña → Mostrar hoja 'Listas'")
print("   2. Agregar nueva empresa en columna B")
print("   3. Actualizar rango: Datos → Validación → cambiar B$3 por B$4, etc.")
print("=" * 60)