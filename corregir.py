from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation

# ============================================================
# CONFIGURACIÓN
# ============================================================
archivo = r"D:\Users\Usuario\Desktop\QuickBooks\data\Facturación Nuevo.xlsx"
MAX_FILAS = 20000

print("=" * 60)
print("🔧 CORRIGIENDO FORMATO Y VALIDACIONES")
print("=" * 60)

# Cargar archivo
print("\n📂 Cargando archivo...")
wb = load_workbook(archivo)
ws = wb.active

# Última columna (X = 24)
ULTIMA_COLUMNA = "X"

print(f"   Rango a aplicar: A2:{ULTIMA_COLUMNA}{MAX_FILAS}")

# ============================================================
# 1. ELIMINAR TODAS LAS VALIDACIONES DE DATOS EXISTENTES
# ============================================================
print("\n🗑️  Eliminando validaciones de datos anteriores...")
ws.data_validations.dataValidation = []

# ============================================================
# 2. ELIMINAR REGLAS DE FORMATO CONDICIONAL DE STATUS
# ============================================================
print("🗑️  Limpiando reglas de formato condicional de STATUS...")

reglas_a_mantener = []
for rango, reglas in list(ws.conditional_formatting._cf_rules.items()):
    for regla in reglas:
        if hasattr(regla, 'formula') and regla.formula:
            formula_str = str(regla.formula)
            # Mantener reglas que NO son de columna A (STATUS)
            if '$A' not in formula_str:
                reglas_a_mantener.append((rango, regla))

ws.conditional_formatting._cf_rules = {}

for rango, regla in reglas_a_mantener:
    try:
        rango_str = str(rango).replace('<ConditionalFormatting ', '').replace('>', '').strip()
        ws.conditional_formatting.add(rango_str, regla)
    except:
        pass

# ============================================================
# 3. CREAR NUEVAS REGLAS DE FORMATO CONDICIONAL
# ============================================================
print("\n🎨 Aplicando formato condicional para STATUS...")

azul_tenue = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
amarillo = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
naranja = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

rango = f"A2:{ULTIMA_COLUMNA}{MAX_FILAS}"

ws.conditional_formatting.add(rango, FormulaRule(formula=['$A2="Ejecutar"'], fill=azul_tenue))
print("   ✅ Ejecutar → Azul tenue")

ws.conditional_formatting.add(rango, FormulaRule(formula=['$A2="Pendiente"'], fill=amarillo))
print("   ✅ Pendiente → Amarillo")

ws.conditional_formatting.add(rango, FormulaRule(formula=['$A2="Revisión"'], fill=naranja))
print("   ✅ Revisión → Naranja")

ws.conditional_formatting.add(rango, FormulaRule(formula=['$A2="Lista"'], fill=verde))
print("   ✅ Lista → Verde")

# ============================================================
# 4. VALIDACIÓN DE STATUS (Columna A) - Usa hoja Listas
# ============================================================
print("\n📋 Aplicando validación STATUS (columna A)...")

dv_status = DataValidation(
    type="list",
    formula1="=Listas!$A$1:$A$4",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Status no válido",
    error="Solo se permiten: Ejecutar, Pendiente, Revisión, Lista"
)
dv_status.add(f"A2:A{MAX_FILAS}")
ws.add_data_validation(dv_status)
print(f"   ✅ A2:A{MAX_FILAS} → Lista desde hoja 'Listas'")

# ============================================================
# 5. VALIDACIÓN DE EMPRESA (Columna F) - Usa hoja Listas
# ============================================================
print("\n🏢 Aplicando validación Empresa (columna F)...")

dv_empresa = DataValidation(
    type="list",
    formula1="=Listas!$B$1:$B$3",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Empresa no válida",
    error="Seleccione una empresa de la lista"
)
dv_empresa.add(f"F2:F{MAX_FILAS}")
ws.add_data_validation(dv_empresa)
print(f"   ✅ F2:F{MAX_FILAS} → Lista desde hoja 'Listas'")

# ============================================================
# 6. VALIDACIÓN DE FECHAS (Columnas K y L)
# ============================================================
print("\n📅 Aplicando validación de fechas (columnas K y L)...")

dv_fecha = DataValidation(
    type="date",
    operator="greaterThan",
    formula1="2020-01-01",
    allow_blank=True,
    showErrorMessage=True,
    errorTitle="Fecha no válida",
    error="Ingrese una fecha válida (después de 01/01/2020). No escriba texto."
)
dv_fecha.add(f"K2:K{MAX_FILAS}")
dv_fecha.add(f"L2:L{MAX_FILAS}")
ws.add_data_validation(dv_fecha)
print(f"   ✅ K2:K{MAX_FILAS} → Solo fechas")
print(f"   ✅ L2:L{MAX_FILAS} → Solo fechas")

# ============================================================
# 7. GUARDAR
# ============================================================
print("\n💾 Guardando archivo...")
wb.save(archivo)

print(f"\n✅ ARCHIVO CORREGIDO: {archivo}")
print("=" * 60)
print("\n📝 RESUMEN COMPLETO:")
print(f"   Rango aplicado: Filas 2 a {MAX_FILAS}")
print("")
print("   🎨 FORMATO CONDICIONAL (toda la fila):")
print("      🔵 Ejecutar  → Azul tenue")
print("      🟡 Pendiente → Amarillo")
print("      🟠 Revisión  → Naranja")
print("      🟢 Lista     → Verde")
print("")
print("   📋 LISTAS DESPLEGABLES:")
print("      Columna A (STATUS)  → Listas!A1:A4")
print("      Columna F (Empresa) → Listas!B1:B3")
print("")
print("   📅 VALIDACIÓN DE FECHAS:")
print("      Columna K (Fecha trabajo)")
print("      Columna L (Fecha reportado)")
print("=" * 60)